from openpyxl import load_workbook

from dataclass.Community import Community


class CommunityExcelGenerator(object):
    FILE_NAME1 = '模板.xlsx'
    FILE_NAME2 = '模板.xls'

    GEN_FILE_PREFIX = '社区_'

    SHEET_NAME_TARGET_SOURCE = '指定数据源'
    SHEET_NAME_SOURCES = '数据源'

    TARGET_COLUMN_NAME = 'A'    #目标数据社区名的列
    TARGET_COLUMN_PROPORTION = 'B'  # 需要改变的户数比例的列
    TARGET_COLUMN_TIME_BEFORE = 'C'  # 需要改变的时间前变量的列
    TARGET_COLUMN_TIME_AFTER = 'D'  # 需要改变的时间后变量的列
    TARGET_DATA_LINE = '2'  # 数据是从第几行开始的。 最小值为1

    SOURCE_COLUMN_NAME = 'A'  # 需要改变的社区名的列
    SOURCE_COLUMN_PROPORTION = 'B'  # 需要改变的户数比例的列
    SOURCE_COLUMN_TIME_BEFORE = 'C'  # 需要改变的时间前变量的列
    SOURCE_COLUMN_TIME_AFTER = 'D'  # 需要改变的时间后变量的列
    SOURCE_DATA_START_LINE = '2'  # 数据是从第几行开始的。 最小值为1

    def __init__(self):
        self.open_workbook = self.open_excel(self.FILE_NAME1)
        if self.open_workbook is None:
            self.open_workbook = self.open_excel(self.FILE_NAME2)

        if self.open_workbook is not None:
            self.communities = self.get_all_communities_data()
            # print(f'communities len = {len(self.communities)}')
            # print(f'communities = {self.communities}')

    def generate_multiple_excels(self):
        for source in self.communities:
            self.change_target_data(source)
            try:
                self.open_workbook.save(f'{self.GEN_FILE_PREFIX}{source.name}.xlsx')
            except Exception as err:
                print (f'{err}')
            else:
                print (f'{source.name} 成功生成！')



    def find_sheet_by_name(self, sheet_name):
        found_sheet = None
        for sheet in self.open_workbook.worksheets:
            if sheet.title == sheet_name:
                found_sheet = sheet

        return found_sheet

    def change_target_data(self, community):
        target_sheet = self.find_sheet_by_name(self.SHEET_NAME_TARGET_SOURCE)

        read_name = target_sheet[self.TARGET_COLUMN_NAME + self.TARGET_DATA_LINE].value
        target_sheet[self.TARGET_COLUMN_NAME + self.TARGET_DATA_LINE] = community.name
        target_sheet[self.TARGET_COLUMN_PROPORTION + self.TARGET_DATA_LINE] = community.proportion
        target_sheet[self.TARGET_COLUMN_TIME_BEFORE + self.TARGET_DATA_LINE] = community.time_before
        target_sheet[self.TARGET_COLUMN_TIME_AFTER + self.TARGET_DATA_LINE] = community.time_after


    def get_all_communities_data(self):
        communities = []
        index = self.SOURCE_DATA_START_LINE

        while True:
            community = self.get_community_data(index)
            if community.name is not None:
                communities.append(community)
            else:
                break

            index = str(int(index) + 1)

        return communities

    def get_community_data(self, target_line):
        community = Community()
        boundExcel = load_workbook(self.FILE_NAME1)
        if boundExcel is None:
            boundExcel = load_workbook(self.FILE_NAME2)

        if (len(boundExcel.worksheets) == 4):
            worksheet = boundExcel.worksheets[3]
            community.name = worksheet[self.SOURCE_COLUMN_NAME + target_line].value
            community.proportion = worksheet[self.SOURCE_COLUMN_PROPORTION + target_line].value
            community.time_before = worksheet[self.SOURCE_COLUMN_TIME_BEFORE + target_line].value
            community.time_after = worksheet[self.SOURCE_COLUMN_TIME_AFTER + target_line].value
        else:
            print('Error, the file not exist')

        return community

    def open_excel(self, file_name):
        open_workbook = None
        try:
            open_workbook = load_workbook(file_name)
        except FileNotFoundError:
            print(f"未找到名为'{file_name}'的文件")
        except Exception as err:
            print(1, err)
        else:
            print(f"已找到名为'{file_name}'的文件")

        return open_workbook

    def changeValue(self, targetValue):
        targetValue
